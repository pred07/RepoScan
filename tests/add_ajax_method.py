import os
from openpyxl.styles import PatternFill, Font

# Read the reporter.py file
reporter_path = r'c:\Users\groot\Music\susu\RepoScan\src\reporter.py'

with open(reporter_path, 'r', encoding='utf-8') as f:
    lines = f.readlines()

# Find the line with the last method (before the end of the class)
# We'll insert before the last line (which should be empty or end of file)

ajax_method = '''
    def _create_ajax_sheet(self):
        """Generates AJAX Code tab with detected AJAX calls and color coding."""
        # Columns: #, File Path, File Name, AJAX Type, Start Line, End Line, Endpoint/URL, Has Server Dependencies, Is Inline, Code Snippet, Full Code
        headers = ["#", "File Path", "File Name", "AJAX Type", "Start Line", "End Line", 
                   "Endpoint/URL", "Has Server Dependencies", "Is Inline", "Code Snippet", "Full Code"]
        data = []
        
        # Filter only AJAX-detected findings
        ajax_findings = [f for f in self.findings if f.ajax_detected]
        for i, f in enumerate(ajax_findings, 1):
            data.append([
                i,
                f.file_path,
                os.path.basename(f.file_path),
                f.ajax_pattern or "Unknown",
                f.start_line,
                f.end_line,
                f.endpoint_url or "Unknown/Dynamic",
                "Yes" if f.has_server_deps else "No",
                "Yes" if f.is_inline_ajax else "No",
                f.snippet,
                f.full_code
            ])
        
        self._create_sheet("AJAX Code", headers, data)
        
        # Apply color coding to the AJAX tab
        if "AJAX Code" in self.wb.sheetnames:
            ws = self.wb["AJAX Code"]
            
            # Color code rows (starting from row 2, after header)
            for row in range(2, ws.max_row + 1):
                # Server Dependencies column (H = column 8)
                server_cell = ws.cell(row=row, column=8)
                if server_cell.value == "Yes":
                    server_cell.fill = PatternFill("solid", fgColor="FFC7CE")  # Red
                    server_cell.font = Font(color="9C0006")
                else:
                    server_cell.fill = PatternFill("solid", fgColor="C6EFCE")  # Green
                    server_cell.font = Font(color="006100")
                
                # Is Inline column (I = column 9)
                inline_cell = ws.cell(row=row, column=9)
                if inline_cell.value == "Yes":
                    inline_cell.fill = PatternFill("solid", fgColor="FFEB9C")  # Yellow
                    inline_cell.font = Font(color="9C6500")
'''

# Append the method
lines.append(ajax_method)

# Write back
with open(reporter_path, 'w', encoding='utf-8') as f:
    f.writelines(lines)

print("✓ AJAX method added to reporter.py")
